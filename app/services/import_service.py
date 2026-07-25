import csv
import io
import re
import unicodedata
from datetime import date, datetime
from urllib.parse import parse_qs, urlparse
from urllib.request import urlopen

from openpyxl import load_workbook

from app import db
from app.models import Employee, User


# =========================================================
# TÊN CỘT ĐƯỢC CHẤP NHẬN
# =========================================================

ALIASES = {
    "full_name": [
        "ho ten",
        "họ tên",
        "hoten",
        "full name",
        "fullname",
        "name",
        "ten nhan vien",
        "tên nhân viên",
    ],

    "department": [
        "phong ban",
        "phòng ban",
        "department",
        "bo phan",
        "bộ phận",
    ],

    "email": [
        "email",
        "e-mail",
        "mail",
    ],

    "phone": [
        "so dien thoai",
        "số điện thoại",
        "dien thoai",
        "điện thoại",
        "phone",
        "telephone",
        "sdt",
    ],

    "position": [
        "chuc vu",
        "chức vụ",
        "position",
        "job title",
    ],

    "hire_date": [
        "ngay vao lam",
        "ngày vào làm",
        "hire date",
        "start date",
    ],

    "skills": [
        "ky nang",
        "kỹ năng",
        "skill",
        "skills",
    ],

    "address": [
        "dia chi",
        "địa chỉ",
        "address",
    ],
}


# =========================================================
# XÓA DẤU TIẾNG VIỆT
# =========================================================

def remove_accents(value):

    value = str(value or "")

    normalized = unicodedata.normalize(
        "NFD",
        value
    )

    value = "".join(
        character
        for character in normalized
        if unicodedata.category(character)
        != "Mn"
    )

    value = value.replace(
        "đ",
        "d"
    )

    value = value.replace(
        "Đ",
        "D"
    )

    return value


# =========================================================
# CHUẨN HÓA HEADER
# =========================================================

def normalize_header(value):

    value = remove_accents(
        value
    ).lower().strip()

    value = re.sub(
        r"[_\-]+",
        " ",
        value
    )

    value = re.sub(
        r"\s+",
        " ",
        value
    )

    return value


# =========================================================
# LÀM SẠCH GIÁ TRỊ
# =========================================================

def clean_value(value):

    if value is None:
        return ""

    if isinstance(
        value,
        float
    ) and value.is_integer():

        return str(
            int(value)
        )

    return str(
        value
    ).strip()


# =========================================================
# ĐỔI TÊN CỘT VỀ FORMAT HỆ THỐNG
# =========================================================

def map_record(record):

    result = {
        "full_name": "",
        "department": "",
        "email": "",
        "phone": "",
        "position": "",
        "hire_date": "",
        "skills": "",
        "address": "",
    }

    normalized_aliases = {}

    for field, aliases in ALIASES.items():

        normalized_aliases[field] = {
            normalize_header(alias)
            for alias in aliases
        }

    for raw_key, raw_value in record.items():

        key = normalize_header(
            raw_key
        )

        for field, aliases in normalized_aliases.items():

            if key in aliases:

                result[field] = clean_value(
                    raw_value
                )

                break

    return result


# =========================================================
# CHUYỂN NGÀY
# =========================================================

def parse_date(value):

    if not value:
        return None

    if isinstance(
        value,
        datetime
    ):
        return value.date()

    if isinstance(
        value,
        date
    ):
        return value

    text = str(
        value
    ).strip()

    formats = [
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%m/%d/%Y",
        "%Y/%m/%d",
    ]

    for date_format in formats:

        try:

            return datetime.strptime(
                text,
                date_format
            ).date()

        except ValueError:
            pass

    return None


# =========================================================
# TẠO USERNAME KHÔNG TRÙNG
# =========================================================

def create_unique_username(
    full_name,
    email=""
):

    # Ưu tiên phần trước @ của email

    if email and "@" in email:

        base = email.split(
            "@"
        )[0]

    else:

        base = remove_accents(
            full_name
        ).lower()

        base = re.sub(
            r"[^a-z0-9]+",
            "",
            base
        )

    if not base:

        base = "employee"

    base = base[:50]

    username = base

    counter = 1

    while (
        User.query
        .filter_by(
            username=username
        )
        .first()
        is not None
    ):

        username = (
            f"{base}{counter}"
        )

        counter += 1

    return username


# =========================================================
# IMPORT RECORDS
# =========================================================

def import_records(
    records,
    create_accounts=False
):

    result = {
        "created": 0,
        "updated": 0,
        "skipped": 0,
        "accounts": [],
    }

    try:

        for raw_record in records:

            record = map_record(
                raw_record
            )

            full_name = (
                record["full_name"]
                .strip()
            )

            email = (
                record["email"]
                .strip()
                .lower()
            )

            # ---------------------------------------------
            # HỌ TÊN + EMAIL BẮT BUỘC
            # ---------------------------------------------

            if (
                not full_name
                or not email
            ):

                result[
                    "skipped"
                ] += 1

                continue

            # ---------------------------------------------
            # TÌM EMPLOYEE THEO EMAIL
            # ---------------------------------------------

            employee = (
                Employee.query
                .filter(
                    db.func.lower(
                        Employee.email
                    )
                    == email
                )
                .first()
            )

            hire_date = parse_date(
                record["hire_date"]
            )

            if employee:

                # -----------------------------------------
                # UPDATE
                # -----------------------------------------

                employee.full_name = (
                    full_name
                )

                employee.department = (
                    record["department"]
                    or employee.department
                    or "Chưa cập nhật"
                )

                employee.phone = (
                    record["phone"]
                    or employee.phone
                )

                employee.position = (
                    record["position"]
                    or employee.position
                )

                employee.skills = (
                    record["skills"]
                    or employee.skills
                )

                employee.address = (
                    record["address"]
                    or employee.address
                )

                if hire_date:

                    employee.hire_date = (
                        hire_date
                    )

                result[
                    "updated"
                ] += 1

            else:

                # -----------------------------------------
                # CREATE
                # -----------------------------------------

                employee = Employee(
                    full_name=full_name,

                    department=(
                        record["department"]
                        or "Chưa cập nhật"
                    ),

                    email=email,

                    phone=(
                        record["phone"]
                        or None
                    ),

                    position=(
                        record["position"]
                        or None
                    ),

                    hire_date=hire_date,

                    skills=(
                        record["skills"]
                        or None
                    ),

                    address=(
                        record["address"]
                        or None
                    ),

                    active=True,
                )

                db.session.add(
                    employee
                )

                db.session.flush()

                result[
                    "created"
                ] += 1

            # ---------------------------------------------
            # TẠO ACCOUNT
            # ---------------------------------------------

            if (
                create_accounts
                and employee.user is None
            ):

                username = (
                    create_unique_username(
                        full_name,
                        email
                    )
                )

                password = "123456"

                user = User(
                    username=username,
                    full_name=full_name,
                    role="employee",
                    employee_id=employee.id,
                )

                user.set_password(
                    password
                )

                db.session.add(
                    user
                )

                db.session.flush()

                result[
                    "accounts"
                ].append(
                    (
                        username,
                        password
                    )
                )

        db.session.commit()

    except Exception:

        db.session.rollback()

        raise

    return result


# =========================================================
# ĐỌC CSV
# =========================================================

def read_csv_file(
    file_path
):

    records = []

    with open(
        file_path,
        "r",
        encoding="utf-8-sig",
        newline=""
    ) as file:

        reader = csv.DictReader(
            file
        )

        for row in reader:

            records.append(
                dict(row)
            )

    return records


# =========================================================
# ĐỌC EXCEL
# =========================================================

def read_xlsx_file(
    file_path
):

    workbook = load_workbook(
        filename=file_path,
        read_only=True,
        data_only=True
    )

    worksheet = workbook.active

    rows = worksheet.iter_rows(
        values_only=True
    )

    try:

        headers = next(
            rows
        )

    except StopIteration:

        workbook.close()

        return []

    headers = [
        clean_value(header)
        for header in headers
    ]

    records = []

    for row in rows:

        if not any(
            value is not None
            and clean_value(value)
            for value in row
        ):
            continue

        record = {}

        for index, header in enumerate(
            headers
        ):

            if not header:
                continue

            value = (
                row[index]
                if index < len(row)
                else ""
            )

            record[header] = value

        records.append(
            record
        )

    workbook.close()

    return records


# =========================================================
# GOOGLE SHEETS -> CSV URL
# =========================================================

def google_sheet_export_url(
    sheet_url
):

    sheet_url = (
        sheet_url
        .strip()
    )

    match = re.search(
        r"/spreadsheets/d/([^/]+)",
        sheet_url
    )

    if not match:

        raise ValueError(
            "Link Google Sheets không hợp lệ."
        )

    sheet_id = match.group(
        1
    )

    parsed = urlparse(
        sheet_url
    )

    query = parse_qs(
        parsed.query
    )

    gid = "0"

    if "gid" in query:

        gid = query[
            "gid"
        ][0]

    else:

        fragment_match = re.search(
            r"gid=(\d+)",
            parsed.fragment
        )

        if fragment_match:

            gid = fragment_match.group(
                1
            )

    return (
        "https://docs.google.com/"
        f"spreadsheets/d/{sheet_id}/"
        "export?format=csv"
        f"&gid={gid}"
    )


# =========================================================
# ĐỌC GOOGLE SHEETS
# =========================================================

def read_google_sheet(
    sheet_url
):

    export_url = (
        google_sheet_export_url(
            sheet_url
        )
    )

    with urlopen(
        export_url,
        timeout=20
    ) as response:

        content = response.read()

    text = content.decode(
        "utf-8-sig"
    )

    reader = csv.DictReader(
        io.StringIO(text)
    )

    records = [
        dict(row)
        for row in reader
    ]

    return records